import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from openpyxl import Workbook, load_workbook


# ------------------ VARIÁVEL GLOBAL ------------------

arquivo_atual = None  # Guarda o caminho da planilha aberta


# ------------------ FUNÇÕES ------------------

def calcular():
    try:
        custo = float(entry_custo.get())
        imposto_percentual = float(entry_imposto.get())
        lucro_percentual = float(entry_lucro.get())

        imposto = custo * (imposto_percentual / 100)
        lucro = custo * (lucro_percentual / 100)
        preco_final = custo + imposto + lucro

        label_resultado.config(text=f"R$ {preco_final:.2f}")

        dados["custo"] = custo
        dados["imposto"] = imposto
        dados["lucro"] = lucro
        dados["preco_final"] = preco_final

    except ValueError:
        messagebox.showerror("Erro", "Preencha todos os campos corretamente.")


def salvar_planilha():
    global arquivo_atual

    try:
        # Se ainda não abriu nenhuma planilha
        if not arquivo_atual:
            arquivo = filedialog.asksaveasfilename(defaultextension=".xlsx")
            if not arquivo:
                return

            wb = Workbook()
            ws = wb.active
            ws.title = "Precificação"

            # Cabeçalho
            ws.append(["Custo", "Imposto", "Lucro", "Preço Final"])

            arquivo_atual = arquivo
        else:
            wb = load_workbook(arquivo_atual)
            ws = wb.active

        # ✅ ADICIONA UM NOVO REGISTRO SEM APAGAR O ANTIGO
        ws.append([
            dados.get("custo", 0),
            dados.get("imposto", 0),
            dados.get("lucro", 0),
            dados.get("preco_final", 0)
        ])

        wb.save(arquivo_atual)
        messagebox.showinfo("Sucesso", "Novo registro adicionado na planilha!")

    except Exception as e:
        messagebox.showerror("Erro", str(e))


def abrir_planilha():
    global arquivo_atual

    try:
        arquivo = filedialog.askopenfilename(filetypes=[("Planilhas", "*.xlsx")])
        if not arquivo:
            return

        wb = load_workbook(arquivo)
        ws = wb.active

        # Carrega o ÚLTIMO registro apenas para edição rápida
        ultima_linha = ws.max_row

        entry_custo.delete(0, tk.END)
        entry_imposto.delete(0, tk.END)
        entry_lucro.delete(0, tk.END)

        entry_custo.insert(0, ws.cell(row=ultima_linha, column=1).value)
        entry_imposto.insert(0, ws.cell(row=ultima_linha, column=2).value)
        entry_lucro.insert(0, ws.cell(row=ultima_linha, column=3).value)

        arquivo_atual = arquivo

        messagebox.showinfo("Sucesso", "Planilha carregada! Novo salvamento será em nova linha.")

    except Exception as e:
        messagebox.showerror("Erro", str(e))


# ------------------ INTERFACE ------------------

app = tk.Tk()
app.title("Sistema de Precificação Profissional")
app.geometry("420x520")
app.configure(bg="#1e1e2e")
app.resizable(False, False)

dados = {}

style = ttk.Style()
style.theme_use("clam")

style.configure("TLabel", background="#1e1e2e", foreground="white", font=("Segoe UI", 11))
style.configure("TEntry", font=("Segoe UI", 11))

# ---------- TÍTULO ----------
titulo = tk.Label(
    app,
    text="Sistema de Precificação",
    font=("Segoe UI", 18, "bold"),
    bg="#1e1e2e",
    fg="#00ffd5"
)
titulo.pack(pady=15)

# ---------- CARD PRINCIPAL ----------
frame = tk.Frame(app, bg="#2a2a3d", padx=20, pady=20)
frame.pack(pady=10)

ttk.Label(frame, text="Custo (R$):").grid(row=0, column=0, sticky="w", pady=8)
entry_custo = ttk.Entry(frame, width=22)
entry_custo.grid(row=0, column=1)

ttk.Label(frame, text="Imposto (%):").grid(row=1, column=0, sticky="w", pady=8)
entry_imposto = ttk.Entry(frame, width=22)
entry_imposto.grid(row=1, column=1)

ttk.Label(frame, text="Lucro (%):").grid(row=2, column=0, sticky="w", pady=8)
entry_lucro = ttk.Entry(frame, width=22)
entry_lucro.grid(row=2, column=1)

# ---------- BOTÃO CALCULAR ----------
btn_calcular = tk.Button(
    frame,
    text="CALCULAR",
    bg="#00ffd5",
    fg="#000",
    font=("Segoe UI", 11, "bold"),
    width=20,
    command=calcular
)
btn_calcular.grid(row=3, column=0, columnspan=2, pady=15)

# ---------- RESULTADO ----------
label_texto = tk.Label(
    frame,
    text="Preço Final",
    font=("Segoe UI", 12),
    bg="#2a2a3d",
    fg="white"
)
label_texto.grid(row=4, column=0, columnspan=2)

label_resultado = tk.Label(
    frame,
    text="R$ 0.00",
    font=("Segoe UI", 18, "bold"),
    bg="#2a2a3d",
    fg="#00ffd5"
)
label_resultado.grid(row=5, column=0, columnspan=2, pady=10)

# ---------- ÁREA DOS BOTÕES ----------
frame_botoes = tk.Frame(app, bg="#1e1e2e")
frame_botoes.pack(pady=20)

btn_salvar = tk.Button(
    frame_botoes,
    text="Adicionar à Planilha",
    bg="#4CAF50",
    fg="white",
    font=("Segoe UI", 11, "bold"),
    width=20,
    command=salvar_planilha
)
btn_salvar.pack(pady=5)

btn_abrir = tk.Button(
    frame_botoes,
    text="Abrir Planilha",
    bg="#ff9800",
    fg="white",
    font=("Segoe UI", 11, "bold"),
    width=20,
    command=abrir_planilha
)
btn_abrir.pack(pady=5)

app.mainloop()
